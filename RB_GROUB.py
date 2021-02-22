import openpyxl

wb_1 = openpyxl.load_workbook('input_wagon.xlsx')
wb_2 = openpyxl.load_workbook('output_wagon.xlsx')

sheet_1 = wb_1['Orders_07']
sheet_2 = wb_1['Orders_08']

sheet_3 = wb_2['KPI']

sheet_4 = wb_1['WagonModelCompatibility']
sheet_5 = wb_1['sources']

summ_1 = 0
summ_2 = 0
summ_3 = 0
summ_4 = 0

k_1 = 0
k_2 = 0
k_3 = 0

rows_1 = sheet_1.max_row
rows_2 = sheet_2.max_row
rows_3 = sheet_4.max_row
rows_4 = sheet_5.max_row

cols_1 = sheet_5.max_column


wagon_model = []

for i in range(3, rows_3 + 1):
    cell_obj_3 = sheet_4.cell(row = i, column = 1)
    wagon_model.append(str(cell_obj_3.value))

for i in range(0, len(wagon_model)):
    print('type '+ str(i) + ': ' + wagon_model[i] + ' -- ', wagon_model.count(wagon_model[i]))

for i in range(3, rows_4 + 1):
    source_m = []
    for j in range(2,5):
        cell = sheet_5.cell(row = i, column = j)
        source_m.append(int(cell.value))
    k_1 = (source_m[0])*((source_m[1])*(source_m[2]))
    k_2 = k_2 + k_1
    k_1 = 0
print(k_2)

for i in range(3, rows_4 + 1):
    m_2 = []
    for j in range(2,4):
        cell = sheet_5.cell(row = i, column = j)
        m_2.append(int(cell.value))
    if ((m_2[0]) > 0):
        k_3 = k_3 + m_2[1]
print(k_3)

for i in range(3, rows_1 + 1):
    cell_obj_1 = sheet_1.cell(row = i, column = 10)
    if (int(cell_obj_1.value) == 1):
        summ_1 = summ_1 + 1
    if (int(cell_obj_1.value) == 2):
        summ_2 = summ_2 + 1
    if (int(cell_obj_1.value) == 3):
        summ_3 = summ_3 + 1
    if (int(cell_obj_1.value) == 4):
        summ_4 = summ_4 + 1
        
for i in range(3, rows_2 + 1):
    cell_obj_2 = sheet_2.cell(row = i, column = 10)
    if (int(cell_obj_2.value) == 1):
        summ_1 = summ_1 + 1
    if (int(cell_obj_2.value) == 2):
        summ_2 = summ_2 + 1
    if (int(cell_obj_2.value) == 3):
        summ_3 = summ_3 + 1
    if (int(cell_obj_2.value) == 4):
        summ_4 = summ_4 + 1
        
sheet_3['B3'] = summ_1
sheet_3['C3'] = summ_2
sheet_3['D3'] = summ_3
sheet_3['E3'] = summ_4
sheet_3['F3'] = summ_1 + summ_2 + summ_3 + summ_4
sheet_3['F12'] = k_2
sheet_3['F11'] = k_3


wb_2.save('output_wagon.xlsx')

