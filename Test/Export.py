import openpyxl
import xlwt
from xlrd import sheet

my_path = "F:\Denis\Ramax\input_wagon2.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_wb_obj.active =0
my_sheet_obj = my_wb_obj.active
my_row = my_sheet_obj.max_row



my_wb = openpyxl.Workbook()
my_wb.active = 0
my_sheet = my_wb.active
myrow = my_sheet.max_row



for i in range(3, my_row + 1):#(3, ..)строка
    cell_obj = my_sheet_obj.cell(row = i, column = 10)#column-столб

    if cell_obj.value == 1:
        print("Own")
        c1 = my_sheet.cell(row=i, column=2)
        c1.value = cell_obj.value

    if cell_obj.value == 2:
        print("Two")
        c2 = my_sheet.cell(row=i, column=3)
        c2.value = cell_obj.value
    if cell_obj.value == 3:
        print("Three")
        c3 = my_sheet.cell(row=i, column=4)
        c3.value = cell_obj.value
    if cell_obj.value == 4:
        print("Four")
        c4 = my_sheet.cell(row=i, column=5)
        c4.value = cell_obj.value






#c = xlwt.Formula('=СЧЁТЗ(B:B)')
#print(c)

cP = my_sheet.cell(row = 2, column = 1)
cP.value = "приоритет"

c1 = my_sheet.cell(row = 2, column = 3)
c1.value = "2"

c1 = my_sheet.cell(row = 2, column = 2)
c1.value = "1"

c1 = my_sheet.cell(row = 2, column = 4)
c1.value = "3"

c1 = my_sheet.cell(row = 2, column = 5)
c1.value = "4"



c1 = my_sheet.cell(row = 4115, column = 2)
c1.value = "Итог от 1"

c1 = my_sheet.cell(row = 4115, column = 3)
c1.value = "Итог от 2"

c1 = my_sheet.cell(row = 4115, column = 4)
c1.value = "Итог от 3"

c1 = my_sheet.cell(row = 4115, column = 5)
c1.value = "Итог от 4"

c1 = my_sheet.cell(row = 4116, column = 2)
c1.value = "=СЧЕТ(B:B)"

c1 = my_sheet.cell(row = 4116, column = 3)
c1.value = "=СЧЕТ(C:C)"

c1 = my_sheet.cell(row = 4116, column = 4)
c1.value = "=СЧЕТ(D:D)"

c1 = my_sheet.cell(row = 4116, column = 5)
c1.value = "=СЧЕТ(E:E)"

#sheet['С2'] = '=СУММ(А1:В2)'

#my_wb.save("F:\Denis\Ramax\output_wagon2.xlsx")

my_wb.save("F:\Denis\Ramax\Chisla.xlsx")

#print(sheet['C2'].value)

